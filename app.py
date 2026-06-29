"""
app.py
Main Flask application — DB Access Provisioning Tool.

Aplikasi ini TIDAK pernah connect/eksekusi apapun ke database target
(Oracle/MySQL/PostgreSQL prod). Aplikasi hanya:
  1. Generate script SQL untuk dicopy manual oleh DBA
  2. Simpan/baca data tracking dari MySQL internal (terpisah, lihat db.py)
  3. Tidak ada login/auth — semua DBA yang punya akses ke tools ini dianggap trusted
"""

import os
from datetime import datetime

from flask import Flask, jsonify, request, render_template
from openpyxl import load_workbook

import db
import scripts

app = Flask(__name__)

ALLOWED_DB_TYPES = {"oracle", "mysql", "postgres"}
ALLOWED_STATUS = {"PENDING", "ACTIVE", "EXTENDED", "LOCKED"}


# ---------------------------------------------------------------------------
# Halaman utama (single page, render template lalu semua interaksi via JS fetch)
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/healthz")
def healthz():
    ok = db.check_connection()
    return jsonify({"db_connected": ok}), (200 if ok else 503)


# ---------------------------------------------------------------------------
# EMPLOYEE MASTER — lookup & administrasi
# ---------------------------------------------------------------------------
@app.route("/api/employees/search")
def search_employees():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])

    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        like = f"%{q}%"
        cur.execute(
            """
            SELECT nik, nama, divisi, unit, subunit
            FROM employee_master
            WHERE nik LIKE %s OR nama LIKE %s
            ORDER BY nama
            LIMIT 8
            """,
            (like, like),
        )
        rows = cur.fetchall()
        return jsonify(rows)
    finally:
        cur.close()
        conn.close()


@app.route("/api/employees", methods=["GET"])
def list_employees():
    search = request.args.get("search", "").strip()
    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        if search:
            like = f"%{search}%"
            cur.execute(
                "SELECT nik, nama, divisi, unit, subunit, last_updated FROM employee_master "
                "WHERE nik LIKE %s OR nama LIKE %s ORDER BY nama",
                (like, like),
            )
        else:
            cur.execute("SELECT nik, nama, divisi, unit, subunit, last_updated FROM employee_master ORDER BY nama")
        rows = cur.fetchall()
        for r in rows:
            if isinstance(r.get("last_updated"), datetime):
                r["last_updated"] = r["last_updated"].strftime("%Y-%m-%d %H:%M:%S")

        # info update terakhir secara keseluruhan (baris yang paling baru diupdate)
        cur.execute("SELECT MAX(last_updated) AS latest FROM employee_master")
        latest_row = cur.fetchone()
        latest = latest_row["latest"] if latest_row else None
        if isinstance(latest, datetime):
            latest = latest.strftime("%Y-%m-%d %H:%M:%S")

        return jsonify({"count": len(rows), "data": rows, "last_updated": latest})
    finally:
        cur.close()
        conn.close()


@app.route("/api/employees/upload-preview", methods=["POST"])
def upload_preview_employees():
    """
    Terima file excel, baca header + beberapa baris pertama,
    kembalikan mentahan agar frontend bisa tampilkan UI mapping kolom.
    File disimpan sementara di memory request, TIDAK ditulis ke disk.
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "File tidak ditemukan"}), 400

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append([("" if c is None else str(c)) for c in row])
            if i >= 500:  # safety limit baris yang dibaca untuk preview
                break
    except Exception as e:
        return jsonify({"error": f"Gagal membaca file: {e}"}), 400

    if len(rows) < 1:
        return jsonify({"error": "File kosong"}), 400

    header = rows[0]
    data_rows = rows[1:]

    return jsonify(
        {
            "header": header,
            "total_rows": len(data_rows),
            "preview_rows": data_rows[:3],
            # seluruh data dikirim balik ke frontend; karena ini internal tool
            # dengan volume kecil (~300 baris), tidak perlu staging di server.
            "all_rows": data_rows,
        }
    )


@app.route("/api/employees/merge", methods=["POST"])
def merge_employees():
    """
    Body JSON:
    {
      "mapping": {"nik": 0, "nama": 1, "divisi": 2, "unit": 3, "subunit": 4},
      "rows": [["198501...", "Budi", "IT", "TI", "Backend"], ...]
    }
    MERGE (UPSERT) berdasarkan NIK:
      - NIK sudah ada -> UPDATE (nama/divisi/unit/subunit + last_updated)
      - NIK baru -> INSERT
      - NIK lama yang tidak ada di file baru -> TETAP ADA, tidak dihapus
    """
    body = request.get_json(force=True)
    mapping = body.get("mapping")
    rows = body.get("rows")

    if not mapping or rows is None:
        return jsonify({"error": "mapping dan rows wajib diisi"}), 400

    required_keys = {"nik", "nama", "divisi", "unit", "subunit"}
    if not required_keys.issubset(mapping.keys()):
        return jsonify({"error": f"mapping harus berisi: {required_keys}"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        conn.start_transaction()
        upsert_sql = """
            INSERT INTO employee_master (nik, nama, divisi, unit, subunit, last_updated)
            VALUES (%s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE
                nama = VALUES(nama),
                divisi = VALUES(divisi),
                unit = VALUES(unit),
                subunit = VALUES(subunit),
                last_updated = NOW()
        """
        values = []
        for row in rows:
            try:
                nik = str(row[mapping["nik"]]).strip()
                nama = str(row[mapping["nama"]]).strip()
                divisi = str(row[mapping["divisi"]]).strip()
                unit = str(row[mapping["unit"]]).strip()
                subunit = str(row[mapping["subunit"]]).strip()
            except IndexError:
                continue
            if not nik:
                continue
            values.append((nik, nama, divisi, unit, subunit))
        if values:
            cur.executemany(upsert_sql, values)
        conn.commit()
        return jsonify({"merged": len(values)})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()



@app.route("/api/clusters/upload-preview", methods=["POST"])
def upload_preview_clusters():
    """
    Terima file excel, baca header + beberapa baris pertama,
    kembalikan mentahan agar frontend bisa tampilkan UI mapping kolom.
    Sama pola dengan upload_preview_employees.
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "File tidak ditemukan"}), 400

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append([("" if c is None else str(c)) for c in row])
            if i >= 500:
                break
    except Exception as e:
        return jsonify({"error": f"Gagal membaca file: {e}"}), 400

    if len(rows) < 1:
        return jsonify({"error": "File kosong"}), 400

    header = rows[0]
    data_rows = rows[1:]

    return jsonify(
        {
            "header": header,
            "total_rows": len(data_rows),
            "preview_rows": data_rows[:3],
            "all_rows": data_rows,
        }
    )


_VALID_RDBMS = {"oracle", "mysql", "postgres"}
_RDBMS_ALIASES = {
    "oracle": "oracle",
    "ora": "oracle",
    "mysql": "mysql",
    "my sql": "mysql",
    "postgres": "postgres",
    "postgresql": "postgres",
    "pg": "postgres",
}


def _normalize_rdbms(raw):
    """Terima variasi penulisan RDBMS dari excel (Oracle/ORACLE/Postgre/PostgreSQL/dll)."""
    key = (raw or "").strip().lower()
    return _RDBMS_ALIASES.get(key)


@app.route("/api/clusters/merge", methods=["POST"])
def merge_clusters():
    """
    Body JSON:
    {
      "mapping": {"rdbms": 0, "cluster_name": 1},
      "rows": [["Oracle", "ora-prod-cluster-01"], ...]
    }
    UPSERT berdasarkan unique (rdbms, cluster_name).
    Baris dengan rdbms yang tidak dikenali (bukan Oracle/MySQL/Postgre) akan di-skip
    dan dilaporkan balik ke frontend.
    """
    body = request.get_json(force=True)
    mapping = body.get("mapping")
    rows = body.get("rows")

    if not mapping or rows is None:
        return jsonify({"error": "mapping dan rows wajib diisi"}), 400

    required_keys = {"rdbms", "cluster_name"}
    if not required_keys.issubset(mapping.keys()):
        return jsonify({"error": f"mapping harus berisi: {required_keys}"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    skipped = []
    try:
        conn.start_transaction()
        upsert_sql = """
            INSERT INTO db_cluster_master (rdbms, cluster_name, created_at)
            VALUES (%s, %s, NOW())
            ON DUPLICATE KEY UPDATE
                cluster_name = VALUES(cluster_name)
        """
        values = []
        for idx, row in enumerate(rows):
            try:
                rdbms_raw = str(row[mapping["rdbms"]]).strip()
                cluster_name = str(row[mapping["cluster_name"]]).strip()
            except IndexError:
                continue
            if not cluster_name:
                continue
            rdbms = _normalize_rdbms(rdbms_raw)
            if not rdbms:
                skipped.append({"row": idx + 2, "rdbms": rdbms_raw, "cluster_name": cluster_name})
                continue
            values.append((rdbms, cluster_name))
        if values:
            cur.executemany(upsert_sql, values)
        conn.commit()
        return jsonify({"merged": len(values), "skipped": skipped})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/clusters", methods=["GET"])
def list_clusters():
    """
    Return daftar cluster + jumlah access_tracking yang memakai (kolom 'used_count'),
    dipakai frontend untuk disable tombol Hapus.
    """
    search = request.args.get("search", "").strip()
    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        base_sql = """
            SELECT c.id, c.rdbms, c.cluster_name, c.created_at,
                   COUNT(t.id) AS used_count
            FROM db_cluster_master c
            LEFT JOIN access_tracking t ON t.db_cluster_id = c.id
        """
        params = []
        if search:
            base_sql += " WHERE c.rdbms LIKE %s OR c.cluster_name LIKE %s"
            like = f"%{search}%"
            params.extend([like, like])
        base_sql += " GROUP BY c.id ORDER BY c.rdbms, c.cluster_name"
        cur.execute(base_sql, params)
        rows = cur.fetchall()
        for r in rows:
            if isinstance(r.get("created_at"), datetime):
                r["created_at"] = r["created_at"].strftime("%Y-%m-%d %H:%M:%S")
        return jsonify(rows)
    finally:
        cur.close()
        conn.close()


@app.route("/api/clusters", methods=["POST"])
def create_cluster():
    """Buat cluster baru secara manual (bukan via excel)."""
    body = request.get_json(force=True)
    rdbms_raw = (body.get("rdbms") or "").strip()
    cluster_name = (body.get("cluster_name") or "").strip()

    rdbms = _normalize_rdbms(rdbms_raw)
    if not rdbms:
        return jsonify({"error": "rdbms harus salah satu dari Oracle, MySQL, Postgres"}), 400
    if not cluster_name:
        return jsonify({"error": "cluster_name wajib diisi"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO db_cluster_master (rdbms, cluster_name, created_at) VALUES (%s, %s, NOW())",
            (rdbms, cluster_name),
        )
        conn.commit()
        return jsonify({"id": cur.lastrowid, "rdbms": rdbms, "cluster_name": cluster_name})
    except Exception as e:
        conn.rollback()
        if "Duplicate" in str(e):
            return jsonify({"error": f"Cluster '{cluster_name}' ({rdbms}) sudah ada"}), 409
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/clusters/<int:cluster_id>", methods=["PUT"])
def update_cluster(cluster_id):
    """Update cluster. Tidak bisa ganti rdbms kalau sudah dipakai (data integrity)."""
    body = request.get_json(force=True)
    rdbms_raw = (body.get("rdbms") or "").strip()
    cluster_name = (body.get("cluster_name") or "").strip()

    rdbms = _normalize_rdbms(rdbms_raw)
    if not rdbms:
        return jsonify({"error": "rdbms harus salah satu dari Oracle, MySQL, Postgres"}), 400
    if not cluster_name:
        return jsonify({"error": "cluster_name wajib diisi"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT COUNT(*) FROM access_tracking WHERE db_cluster_id = %s", (cluster_id,)
        )
        used_count = cur.fetchone()[0]
        cur.execute("SELECT rdbms FROM db_cluster_master WHERE id = %s", (cluster_id,))
        existing = cur.fetchone()
        if not existing:
            return jsonify({"error": "Cluster tidak ditemukan"}), 404
        if used_count > 0 and existing[0] != rdbms:
            return jsonify({"error": "Tidak bisa ubah RDBMS, cluster ini masih dipakai di tracking"}), 409

        cur.execute(
            "UPDATE db_cluster_master SET rdbms=%s, cluster_name=%s WHERE id=%s",
            (rdbms, cluster_name, cluster_id),
        )
        conn.commit()
        return jsonify({"id": cluster_id, "updated": True})
    except Exception as e:
        conn.rollback()
        if "Duplicate" in str(e):
            return jsonify({"error": f"Cluster '{cluster_name}' ({rdbms}) sudah ada"}), 409
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/clusters/<int:cluster_id>", methods=["DELETE"])
def delete_cluster(cluster_id):
    """
    Hapus cluster. Block kalau masih dipakai di access_tracking
    (FK ON DELETE RESTRICT akan throw error, kita tangkap jadi pesan friendly).
    """
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM db_cluster_master WHERE id = %s", (cluster_id,))
        conn.commit()
        return jsonify({"id": cluster_id, "deleted": True})
    except Exception as e:
        conn.rollback()
        if "foreign key constraint" in str(e).lower() or "1451" in str(e):
            return jsonify({"error": "Cluster tidak bisa dihapus, masih dipakai di Active Access Dashboard"}), 409
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/role-suggestion", methods=["POST"])
def role_suggestion():
    """
    Body JSON: {"divisi": "...", "unit": "..."}
    Return: {"divisi_code": "ABC", "unit_code": "DEF", "suggested_base": "ABC_DEF"}

    Kode di-generate sekali per nama (divisi/unit), lalu disimpan permanen
    di role_code_mapping supaya konsisten dan tidak collision dengan nama lain.
    """
    body = request.get_json(force=True)
    divisi = (body.get("divisi") or "").strip()
    unit = (body.get("unit") or "").strip()

    if not divisi or not unit:
        return jsonify({"error": "divisi dan unit wajib diisi"}), 400

    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        def get_or_create_code(name):
            cur.execute("SELECT code FROM role_code_mapping WHERE name = %s", (name,))
            row = cur.fetchone()
            if row:
                return row["code"]

            cur.execute("SELECT code FROM role_code_mapping")
            existing_codes = {r["code"] for r in cur.fetchall()}
            new_code = scripts.pick_unique_role_code(name, existing_codes)

            cur.execute(
                "INSERT INTO role_code_mapping (name, code) VALUES (%s, %s)",
                (name, new_code),
            )
            conn.commit()
            return new_code

        divisi_code = get_or_create_code(divisi)
        unit_code = get_or_create_code(unit)

        return jsonify(
            {
                "divisi_code": divisi_code,
                "unit_code": unit_code,
                "suggested_base": f"{divisi_code}_{unit_code}",
            }
        )
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ---------------------------------------------------------------------------
# PROVISIONING
# ---------------------------------------------------------------------------
@app.route("/api/provisioning/generate-script", methods=["POST"])
def generate_provisioning_script():
    """
    Generate script + password TANPA simpan ke DB.
    DBA preview dulu sebelum confirm-save.
    """
    body = request.get_json(force=True)

    required_fields = [
        "requester", "nik", "unit", "subunit", "created_by",
        "dbtype", "cluster_id", "username", "role", "quarter", "year",
    ]
    missing = [f for f in required_fields if not str(body.get(f, "")).strip()]
    if missing:
        return jsonify({"error": f"Field wajib belum diisi: {missing}"}), 400

    dbtype = body["dbtype"]
    if dbtype not in ALLOWED_DB_TYPES:
        return jsonify({"error": "dbtype tidak valid"}), 400

    if dbtype == "oracle" and not str(body.get("oracle_service_name", "")).strip():
        return jsonify({"error": "Service Name wajib diisi untuk Oracle"}), 400

    if dbtype in ("mysql", "postgres") and not str(body.get("allowlist", "")).strip():
        return jsonify({"error": "Host allowlist wajib diisi untuk MySQL/PostgreSQL"}), 400

    quarter = body["quarter"]
    year = body["year"]
    if quarter not in ("Q1", "Q2", "Q3", "Q4") or not year.isdigit() or len(year) != 4:
        return jsonify({"error": "Quarter/Tahun tidak valid"}), 400

    expiry_display, expiry_iso = scripts.quarter_end_date(quarter, year)
    password = scripts.generate_passphrase()

    script_data = {
        "dbtype": dbtype,
        "username": body["username"].strip(),
        "role": body.get("role", "").strip(),
        "allowlist": body.get("allowlist", "").strip(),
        "expiry_iso": expiry_iso,
    }
    script_text = scripts.build_provisioning_script(script_data, password)

    return jsonify(
        {
            "password": password,
            "script": script_text,
            "expiry_display": expiry_display,
            "expiry_iso": expiry_iso,
        }
    )


@app.route("/api/provisioning/save", methods=["POST"])
def save_provisioning():
    """
    Simpan record ke access_tracking dengan status PENDING.
    Dipanggil setelah DBA klik "Confirm & Save" di preview screen.
    Password TIDAK disimpan di sini (sesuai keputusan: hanya ditampilkan di screen).
    """
    body = request.get_json(force=True)

    required_fields = [
        "requester", "nik", "divisi", "unit", "subunit", "created_by",
        "dbtype", "cluster_id", "username", "role", "expiry_iso",
    ]
    missing = [f for f in required_fields if not str(body.get(f, "")).strip()]
    if missing:
        return jsonify({"error": f"Field wajib belum diisi: {missing}"}), 400

    dbtype = body["dbtype"]
    if dbtype not in ALLOWED_DB_TYPES:
        return jsonify({"error": "dbtype tidak valid"}), 400

    if dbtype == "oracle" and not str(body.get("oracle_service_name", "")).strip():
        return jsonify({"error": "Service Name wajib diisi untuk Oracle"}), 400

    cluster_id = body["cluster_id"]

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        # Ambil snapshot nama cluster untuk display (db_host legacy text column)
        cur.execute(
            "SELECT rdbms, cluster_name FROM db_cluster_master WHERE id = %s", (cluster_id,)
        )
        cluster_row = cur.fetchone()
        if not cluster_row:
            return jsonify({"error": "Cluster tidak ditemukan"}), 404
        cluster_rdbms, cluster_name = cluster_row
        if cluster_rdbms != dbtype:
            return jsonify({"error": "dbtype tidak sesuai dengan RDBMS cluster yang dipilih"}), 400
        host_snapshot = f"{cluster_rdbms} - {cluster_name}"

        cur.execute(
            """
            INSERT INTO access_tracking
                (username, nik, requester, divisi, unit, subunit, db_type, db_host,
                 db_cluster_id, oracle_service_name,
                 role_name, host_allowlist, created_at, expiry_at, status, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'PENDING', %s)
            """,
            (
                body["username"].strip(),
                body["nik"].strip(),
                body["requester"].strip(),
                body["divisi"].strip(),
                body["unit"].strip(),
                body["subunit"].strip(),
                dbtype,
                host_snapshot,
                cluster_id,
                body.get("oracle_service_name", "").strip() or None,
                body["role"].strip(),
                body.get("allowlist", "").strip() or None,
                datetime.now(),
                body["expiry_iso"],
                body["created_by"].strip(),
            ),
        )
        conn.commit()
        new_id = cur.lastrowid

        # Simpan role ke role_master kalau belum ada
        role_name = body["role"].strip()
        cur.execute("SELECT id FROM role_master WHERE role_name = %s", (role_name,))
        if not cur.fetchone():
            try:
                cur.execute(
                    """
                    INSERT INTO role_master (role_name, divisi, unit, suffix, created_at)
                    VALUES (%s, %s, %s, %s, NOW())
                    """,
                    (
                        role_name,
                        body["divisi"].strip(),
                        body["unit"].strip(),
                        body.get("suffix", "RO").strip(),
                    ),
                )
                conn.commit()
            except Exception:
                pass  # Kalau gagal simpan, tidak masalah — role tetap jadi

        return jsonify({"id": new_id, "status": "PENDING"})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------
@app.route("/api/tracking", methods=["GET"])
def list_tracking():
    status_filter = request.args.get("status", "").strip()
    search = request.args.get("search", "").strip()

    query = "SELECT * FROM access_tracking WHERE 1=1"
    params = []

    if status_filter and status_filter in ALLOWED_STATUS:
        query += " AND status = %s"
        params.append(status_filter)

    if search:
        query += " AND (username LIKE %s OR requester LIKE %s)"
        like = f"%{search}%"
        params += [like, like]

    query += " ORDER BY id DESC"

    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(query, params)
        rows = cur.fetchall()
        for r in rows:
            if isinstance(r.get("created_at"), datetime):
                r["created_at"] = r["created_at"].strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(r.get("expiry_at"), datetime):
                r["expiry_at"] = r["expiry_at"].strftime("%Y-%m-%d %H:%M:%S")
        return jsonify(rows)
    finally:
        cur.close()
        conn.close()


def _get_tracking_row(tracking_id):
    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM access_tracking WHERE id = %s", (tracking_id,))
        return cur.fetchone()
    finally:
        cur.close()
        conn.close()


@app.route("/api/tracking/<int:tracking_id>/mark-executed", methods=["POST"])
def mark_executed(tracking_id):
    record = _get_tracking_row(tracking_id)
    if not record:
        return jsonify({"error": "Record tidak ditemukan"}), 404
    if record["status"] != "PENDING":
        return jsonify({"error": "Hanya record berstatus PENDING yang bisa di-mark executed"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE access_tracking SET status = 'ACTIVE' WHERE id = %s", (tracking_id,))
        conn.commit()
        return jsonify({"id": tracking_id, "status": "ACTIVE"})
    finally:
        cur.close()
        conn.close()


@app.route("/api/tracking/<int:tracking_id>/lock-script", methods=["GET"])
def get_lock_script(tracking_id):
    record = _get_tracking_row(tracking_id)
    if not record:
        return jsonify({"error": "Record tidak ditemukan"}), 404
    script_text = scripts.build_lock_script(record)
    return jsonify({"script": script_text})


@app.route("/api/tracking/<int:tracking_id>/mark-locked", methods=["POST"])
def mark_locked(tracking_id):
    record = _get_tracking_row(tracking_id)
    if not record:
        return jsonify({"error": "Record tidak ditemukan"}), 404
    if record["status"] not in ("ACTIVE", "EXTENDED"):
        return jsonify({"error": "Hanya record ACTIVE/EXTENDED yang bisa di-lock"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE access_tracking SET status = 'LOCKED' WHERE id = %s", (tracking_id,))
        conn.commit()
        return jsonify({"id": tracking_id, "status": "LOCKED"})
    finally:
        cur.close()
        conn.close()


@app.route("/api/tracking/<int:tracking_id>/extend-script", methods=["POST"])
def get_extend_script(tracking_id):
    """
    Body JSON: {"quarter": "Q4", "year": "2026"}
    Return script (atau null kalau tidak ada yang perlu dijalankan) + expiry baru.
    """
    record = _get_tracking_row(tracking_id)
    if not record:
        return jsonify({"error": "Record tidak ditemukan"}), 404

    body = request.get_json(force=True)
    quarter = body.get("quarter")
    year = body.get("year")
    if quarter not in ("Q1", "Q2", "Q3", "Q4") or not year or not str(year).isdigit() or len(str(year)) != 4:
        return jsonify({"error": "Quarter/Tahun tidak valid"}), 400

    expiry_display, expiry_iso = scripts.quarter_end_date(quarter, year)
    script_text = scripts.build_extend_script(record, expiry_iso)

    return jsonify(
        {
            "script": script_text,  # null kalau tidak perlu dijalankan
            "expiry_display": expiry_display,
            "expiry_iso": expiry_iso,
            "was_locked": record["status"] == "LOCKED",
        }
    )


@app.route("/api/tracking/<int:tracking_id>/confirm-extend", methods=["POST"])
def confirm_extend(tracking_id):
    """
    Body JSON: {"expiry_iso": "2026-12-31 23:59:59"}
    Update expiry_at dan status -> EXTENDED.
    """
    record = _get_tracking_row(tracking_id)
    if not record:
        return jsonify({"error": "Record tidak ditemukan"}), 404

    body = request.get_json(force=True)
    expiry_iso = body.get("expiry_iso")
    if not expiry_iso:
        return jsonify({"error": "expiry_iso wajib diisi"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE access_tracking SET expiry_at = %s, status = 'EXTENDED' WHERE id = %s",
            (expiry_iso, tracking_id),
        )
        conn.commit()
        return jsonify({"id": tracking_id, "status": "EXTENDED", "expiry_at": expiry_iso})
    finally:
        cur.close()
        conn.close()


@app.route("/api/tracking/<int:tracking_id>", methods=["DELETE"])
def delete_tracking(tracking_id):
    """Hapus record dari access_tracking."""
    record = _get_tracking_row(tracking_id)
    if not record:
        return jsonify({"error": "Record tidak ditemukan"}), 404

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM access_tracking WHERE id = %s", (tracking_id,))
        conn.commit()
        return jsonify({"id": tracking_id, "deleted": True})
    finally:
        cur.close()
        conn.close()


# ---------------------------------------------------------------------------
# ROLE MANAGEMENT
# ---------------------------------------------------------------------------
@app.route("/api/roles", methods=["GET"])
def list_roles():
    """Return daftar role yang sudah pernah dibuat (reusable)."""
    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        search = request.args.get("search", "").strip()
        if search:
            like = f"%{search}%"
            cur.execute(
                "SELECT id, role_name, divisi, unit, suffix, created_at FROM role_master "
                "WHERE role_name LIKE %s OR divisi LIKE %s OR unit LIKE %s ORDER BY role_name",
                (like, like, like),
            )
        else:
            cur.execute("SELECT id, role_name, divisi, unit, suffix, created_at FROM role_master ORDER BY role_name")
        rows = cur.fetchall()
        for r in rows:
            if isinstance(r.get("created_at"), datetime):
                r["created_at"] = r["created_at"].strftime("%Y-%m-%d %H:%M:%S")
        return jsonify(rows)
    finally:
        cur.close()
        conn.close()


@app.route("/api/roles", methods=["POST"])
def create_role():
    """Buat role baru di role_master."""
    body = request.get_json(force=True)
    role_name = (body.get("role_name") or "").strip()
    divisi = (body.get("divisi") or "").strip()
    unit = (body.get("unit") or "").strip()
    suffix = (body.get("suffix") or "RO").strip()

    if not role_name or not divisi or not unit:
        return jsonify({"error": "role_name, divisi, dan unit wajib diisi"}), 400
    if suffix not in ("RO", "RW"):
        return jsonify({"error": "suffix harus RO atau RW"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO role_master (role_name, divisi, unit, suffix, created_at)
            VALUES (%s, %s, %s, %s, NOW())
            """,
            (role_name, divisi, unit, suffix),
        )
        conn.commit()
        return jsonify({"id": cur.lastrowid, "role_name": role_name})
    except Exception as e:
        conn.rollback()
        if "Duplicate" in str(e):
            return jsonify({"error": f"Role '{role_name}' sudah ada"}), 409
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/roles/<int:role_id>", methods=["PUT"])
def update_role(role_id):
    """Update role di role_master."""
    body = request.get_json(force=True)
    role_name = (body.get("role_name") or "").strip()
    divisi = (body.get("divisi") or "").strip()
    unit = (body.get("unit") or "").strip()
    suffix = (body.get("suffix") or "RO").strip()

    if not role_name or not divisi or not unit:
        return jsonify({"error": "role_name, divisi, dan unit wajib diisi"}), 400

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE role_master SET role_name=%s, divisi=%s, unit=%s, suffix=%s WHERE id=%s",
            (role_name, divisi, unit, suffix, role_id),
        )
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({"error": "Role tidak ditemukan"}), 404
        return jsonify({"id": role_id, "updated": True})
    except Exception as e:
        conn.rollback()
        if "Duplicate" in str(e):
            return jsonify({"error": f"Role '{role_name}' sudah ada"}), 409
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/roles/<int:role_id>", methods=["DELETE"])
def delete_role(role_id):
    """Hapus role dari role_master."""
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM role_master WHERE id = %s", (role_id,))
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({"error": "Role tidak ditemukan"}), 404
        return jsonify({"id": role_id, "deleted": True})
    finally:
        cur.close()
        conn.close()


# ---------------------------------------------------------------------------
# STARTUP
# ---------------------------------------------------------------------------
def initialize():
    """Dipanggil sekali saat container start — pastikan schema ada."""
    try:
        db.init_schema()
        print("[startup] Schema OK")
    except Exception as e:
        print(f"[startup] WARNING: gagal inisialisasi schema: {e}")
        print("[startup] Aplikasi tetap jalan, tapi pastikan DB sudah benar sebelum digunakan.")


initialize()

if __name__ == "__main__":
    port = int(os.environ.get("APP_PORT", "5010"))
    app.run(host="0.0.0.0", port=port, debug=False)
